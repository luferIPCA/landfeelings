#!/usr/bin/env python3
"""
gerar_excel.py — LandFeelings
Lê todos os index.json dos catálogos e gera um Excel para preenchimento de metadados.

Uso:
  python gerar_excel.py

Gera: catalogo_metadados.xlsx
Depois de preencher o Excel, correr: python excel_para_json.py
"""

import os
import json
import re

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("A instalar openpyxl...")
    import subprocess, sys
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "--user", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

BASE       = os.path.join(os.path.dirname(__file__), 'images', 'Catalogos')
OUTPUT     = os.path.join(os.path.dirname(__file__), 'catalogo_metadados.xlsx')
EXTENSOES  = {'.jpg', '.jpeg', '.png', '.webp'}

HASHTAGS_VALIDAS = [
    'cerâmica-única', 'cerâmica-molde', 'boneca', 'casa',
    'arte', 'produto-alimentar', 'cestaria', 'outros'
]

# Cores
COR_CAB_FUNDO  = '1B4B9C'  # azul cobalt
COR_CAB_TEXTO  = 'FFFFFF'
COR_ID_FUNDO   = 'F0F4FF'
COR_ANO_FUNDO  = 'FFF8E8'
COR_ALT_FUNDO  = 'F9F9F9'
COR_BORDA      = 'CCCCCC'

borda_lado = Side(style='thin', color=COR_BORDA)
borda_all  = Border(left=borda_lado, right=borda_lado, top=borda_lado, bottom=borda_lado)

COLUNAS = [
    # (header, campo_json, largura, bloqueada, descricao)
    ('ID',           'id',        18, True,  'Identificador único — NÃO EDITAR'),
    ('Ficheiro',     'ficheiro',  28, True,  'Nome do ficheiro — NÃO EDITAR'),
    ('Ano',          'ano',       8,  True,  'Ano do catálogo — NÃO EDITAR'),
    ('Título',       'titulo',    30, False, 'Nome da peça (vazio = nome do ficheiro)'),
    ('Descrição',    'descricao', 45, False, 'Texto descritivo (aceita HTML)'),
    ('Autor',        'autor',     20, False, 'Artesã/autor da peça'),
    ('Preço',        'preco',     12, False, 'Ex: 18,00 € (vazio = Sob consulta)'),
    ('Altura (cm)',  'altura',    12, False, 'Altura em centímetros'),
    ('Largura (cm)', 'largura',   12, False, 'Largura em centímetros'),
    ('Peso (g)',     'peso',      10, False, 'Peso em gramas'),
    ('HashTags',     'hashtags',  40, False, 'Separadas por vírgula: ' + ', '.join(HASHTAGS_VALIDAS)),
]

def gerar_id(ano, n):
    return f'LF-{ano}-{n:03d}'

def carregar_dados():
    """Carrega todos os index.json e devolve lista de peças."""
    pecas = []
    if not os.path.exists(BASE):
        print(f'Pasta não encontrada: {BASE}')
        return pecas

    anos = sorted([d for d in os.listdir(BASE) if os.path.isdir(os.path.join(BASE, d))])
    for ano in anos:
        pasta = os.path.join(BASE, ano)
        index_path = os.path.join(pasta, 'index.json')
        if not os.path.exists(index_path):
            continue
        try:
            with open(index_path, 'r', encoding='utf-8') as f:
                dados = json.load(f)
        except Exception as e:
            print(f'  Erro ao ler {index_path}: {e}')
            continue

        for i, peca in enumerate(dados, 1):
            # Garantir ID único
            if not peca.get('id'):
                peca['id'] = gerar_id(ano, i)

            pecas.append({
                'id':       peca.get('id', gerar_id(ano, i)),
                'ficheiro': peca.get('ficheiro', ''),
                'ano':      ano,
                'titulo':   peca.get('titulo', ''),
                'descricao':peca.get('descricao', ''),
                'autor':    peca.get('autor', ''),
                'data':     peca.get('data', ''),
                'preco':    peca.get('preco', ''),
                'altura':   peca.get('altura', ''),
                'largura':  peca.get('largura', ''),
                'peso':     peca.get('peso', ''),
                'hashtags': ', '.join(peca.get('hashtags', [])) if isinstance(peca.get('hashtags'), list) else peca.get('hashtags', ''),
            })
        print(f'  {ano}: {len(dados)} peças carregadas')

    return pecas

def gerar_excel(pecas):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Catálogo LandFeelings'

    # ── Cabeçalho ──────────────────────────────────────────────────────────
    ws.freeze_panes = 'D2'  # Fixar ID, Ficheiro, Ano e cabeçalho

    for col_idx, (header, campo, largura, bloqueada, desc) in enumerate(COLUNAS, 1):
        cel = ws.cell(row=1, column=col_idx, value=header)
        cel.font      = Font(bold=True, color=COR_CAB_TEXTO, name='Arial', size=10)
        cel.fill      = PatternFill('solid', fgColor=COR_CAB_FUNDO)
        cel.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cel.border    = borda_all
        cel.comment   = None
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    ws.row_dimensions[1].height = 30

    # ── Dados ──────────────────────────────────────────────────────────────
    ano_atual = None
    for row_idx, peca in enumerate(pecas, 2):
        eh_alt = row_idx % 2 == 0
        ano    = peca['ano']
        novo_ano = ano != ano_atual
        ano_atual = ano

        for col_idx, (header, campo, largura, bloqueada, desc) in enumerate(COLUNAS, 1):
            valor = peca.get(campo, '')
            cel   = ws.cell(row=row_idx, column=col_idx, value=valor)
            cel.border    = borda_all
            cel.font      = Font(name='Arial', size=10, color='444444' if bloqueada else '000000')
            cel.alignment = Alignment(vertical='center', wrap_text=col_idx >= 5)

            if bloqueada:
                cel.fill = PatternFill('solid', fgColor=COR_ID_FUNDO if campo == 'id' else
                                       COR_ANO_FUNDO if campo == 'ano' else 'F5F5F5')
            elif eh_alt:
                cel.fill = PatternFill('solid', fgColor=COR_ALT_FUNDO)

            if novo_ano and col_idx == 1:
                ws.row_dimensions[row_idx].height = 18

    # ── Instruções numa 2ª aba ─────────────────────────────────────────────
    wi = wb.create_sheet('Instruções')
    instrucoes = [
        ('COMO USAR ESTE FICHEIRO', True, '1B4B9C'),
        ('', False, '000000'),
        ('1. Não editar as colunas ID, Ficheiro e Ano (fundo azul/amarelo)', False, '000000'),
        ('2. Preencher os campos de metadados nas restantes colunas', False, '000000'),
        ('3. Gravar o ficheiro Excel', False, '000000'),
        ('4. Correr o script: python excel_para_json.py', False, '000000'),
        ('   → O script actualiza os index.json com os metadados', False, '000000'),
        ('', False, '000000'),
        ('CAMPO PREÇO', True, 'E8352A'),
        ('  Formato aceite: 18,00 € ou 18€ ou 18.00', False, '000000'),
        ('  Se vazio: o site mostra "Sob consulta"', False, '000000'),
        ('', False, '000000'),
        ('CAMPO HASHTAGS', True, 'E8352A'),
        ('  Valores válidos (separados por vírgula):', False, '000000'),
    ] + [(f'  • {h}', False, '555555') for h in HASHTAGS_VALIDAS] + [
        ('', False, '000000'),
        ('ID ÚNICO', True, 'E8352A'),
        ('  Formato: LF-ANO-NNN (ex: LF-2022-001)', False, '000000'),
        ('  Atribuído automaticamente pelo script gerar_catalogos.py', False, '000000'),
        ('  Nunca alterar manualmente', False, '000000'),
    ]
    wi.column_dimensions['A'].width = 60
    for i, (texto, bold, cor) in enumerate(instrucoes, 1):
        c = wi.cell(row=i, column=1, value=texto)
        c.font = Font(name='Arial', size=11, bold=bold, color=cor)

    wb.save(OUTPUT)
    print(f'\nExcel gerado: {OUTPUT}')
    print(f'Total de peças: {len(pecas)}')

def main():
    print('LandFeelings — Gerador de Excel de Metadados')
    print(f'Base: {BASE}\n')
    pecas = carregar_dados()
    if not pecas:
        print('Nenhuma peça encontrada. Verifica se os index.json existem.')
        return
    gerar_excel(pecas)
    print('\nPreenche o Excel e corre: python excel_para_json.py')

if __name__ == '__main__':
    main()
