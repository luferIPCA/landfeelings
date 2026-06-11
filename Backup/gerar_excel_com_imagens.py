#!/usr/bin/env python3
"""
gerar_excel_com_imagens.py — LandFeelings
Igual ao gerar_excel.py mas com miniaturas de cada peça (60x60px).

Uso:
  python gerar_excel_com_imagens.py

Gera: catalogo_metadados_imagens.xlsx
AVISO: ficheiro mais pesado (~10-30MB dependendo do número de peças)
"""

import os
import json
import io
import sys

# Instalar dependências se necessário
for modulo, pacote in [('openpyxl', 'openpyxl'), ('PIL', 'Pillow')]:
    try:
        __import__(modulo)
    except ImportError:
        print(f'A instalar {pacote}...')
        import subprocess
        subprocess.run([sys.executable, '-m', 'pip', 'install', pacote, '--user', '-q'])

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

BASE   = os.path.join(os.path.dirname(__file__), 'images', 'Catalogos')
OUTPUT = os.path.join(os.path.dirname(__file__), 'catalogo_metadados_imagens.xlsx')

MINIATURA_SIZE = (70, 70)   # pixels
ALTURA_LINHA   = 58         # pontos Excel (≈ 70px)

COR_CAB_FUNDO = '1B4B9C'
COR_CAB_TEXTO = 'FFFFFF'
COR_ID_FUNDO  = 'F0F4FF'
COR_ANO_FUNDO = 'FFF8E8'
COR_ALT_FUNDO = 'F9F9F9'
COR_BORDA     = 'CCCCCC'

borda_lado = Side(style='thin', color=COR_BORDA)
bordas     = Border(left=borda_lado, right=borda_lado, top=borda_lado, bottom=borda_lado)
margens    = {'top': 80, 'bottom': 80, 'left': 120, 'right': 120}

# Coluna de imagem + restantes
COLUNAS = [
    ('Img',          'img',       10, True,  'Miniatura da peça'),
    ('ID',           'id',        18, True,  'Identificador único — NÃO EDITAR'),
    ('Ficheiro',     'ficheiro',  25, True,  'Nome do ficheiro — NÃO EDITAR'),
    ('Ano',          'ano',       8,  True,  'Ano do catálogo — NÃO EDITAR'),
    ('Título',       'titulo',    28, False, 'Nome da peça'),
    ('Descrição',    'descricao', 40, False, 'Texto descritivo'),
    ('Autor',        'autor',     18, False, 'Artesã/autor'),
    ('Preço',        'preco',     12, False, 'Ex: 18,00 €'),
    ('Altura (cm)',  'altura',    11, False, 'Altura em cm'),
    ('Largura (cm)', 'largura',   11, False, 'Largura em cm'),
    ('Peso (g)',     'peso',      10, False, 'Peso em gramas'),
    ('HashTags',     'hashtags',  38, False, 'Categorias separadas por vírgula'),
]

def miniatura_bytes(img_path):
    """Redimensiona imagem para miniatura e devolve bytes PNG."""
    try:
        with PILImage.open(img_path) as img:
            img = img.convert('RGB')
            img.thumbnail(MINIATURA_SIZE, PILImage.LANCZOS)
            # Criar fundo branco e centrar
            fundo = PILImage.new('RGB', MINIATURA_SIZE, (255, 255, 255))
            offset = ((MINIATURA_SIZE[0] - img.width) // 2,
                      (MINIATURA_SIZE[1] - img.height) // 2)
            fundo.paste(img, offset)
            buf = io.BytesIO()
            fundo.save(buf, format='PNG', optimize=True)
            buf.seek(0)
            return buf
    except Exception as e:
        return None

def carregar_dados():
    pecas = []
    if not os.path.exists(BASE):
        print(f'Pasta não encontrada: {BASE}')
        return pecas

    anos = sorted([d for d in os.listdir(BASE) if os.path.isdir(os.path.join(BASE, d))])
    for ano in anos:
        index_path = os.path.join(BASE, ano, 'index.json')
        if not os.path.exists(index_path):
            continue
        try:
            with open(index_path, 'r', encoding='utf-8') as f:
                dados = json.load(f)
        except Exception as e:
            print(f'  Erro ao ler {index_path}: {e}')
            continue

        for i, peca in enumerate(dados, 1):
            img_path = os.path.join(BASE, ano, peca.get('ficheiro', ''))
            pecas.append({
                'id':        peca.get('id', f'LF-{ano}-{i:03d}'),
                'ficheiro':  peca.get('ficheiro', ''),
                'ano':       ano,
                'titulo':    peca.get('titulo', ''),
                'descricao': peca.get('descricao', ''),
                'autor':     peca.get('autor', ''),
                'data':      peca.get('data', ''),
                'preco':     peca.get('preco', ''),
                'altura':    peca.get('altura', ''),
                'largura':   peca.get('largura', ''),
                'peso':      peca.get('peso', ''),
                'hashtags':  ', '.join(peca.get('hashtags', [])) if isinstance(peca.get('hashtags'), list) else peca.get('hashtags', ''),
                'img_path':  img_path if os.path.exists(img_path) else None,
            })
        print(f'  {ano}: {len(dados)} peças')

    return pecas

def gerar_excel(pecas):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Catálogo LandFeelings'
    ws.freeze_panes = 'E2'  # fixar imagem + ID + Ficheiro + Ano + cabeçalho

    # ── Cabeçalho ──────────────────────────────────────────────────────────
    for col_idx, (header, campo, largura, bloqueada, desc) in enumerate(COLUNAS, 1):
        cel = ws.cell(row=1, column=col_idx, value=header)
        cel.font      = Font(bold=True, color=COR_CAB_TEXTO, name='Arial', size=10)
        cel.fill      = PatternFill('solid', fgColor=COR_CAB_FUNDO)
        cel.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cel.border    = bordas
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    ws.row_dimensions[1].height = 28

    total = len(pecas)
    print(f'\nA gerar Excel com {total} peças...')

    for row_idx, peca in enumerate(pecas, 2):
        ws.row_dimensions[row_idx].height = ALTURA_LINHA
        eh_alt = row_idx % 2 == 0

        for col_idx, (header, campo, largura, bloqueada, desc) in enumerate(COLUNAS, 1):
            if campo == 'img':
                continue  # imagem inserida separadamente
            valor = peca.get(campo, '')
            cel   = ws.cell(row=row_idx, column=col_idx, value=valor)
            cel.border    = bordas
            cel.font      = Font(name='Arial', size=10, color='555555' if bloqueada else '000000')
            cel.alignment = Alignment(vertical='center', wrap_text=col_idx >= 5)
            if bloqueada:
                cor = COR_ID_FUNDO if campo == 'id' else COR_ANO_FUNDO if campo == 'ano' else 'F0F0F0'
                cel.fill = PatternFill('solid', fgColor=cor)
            elif eh_alt:
                cel.fill = PatternFill('solid', fgColor=COR_ALT_FUNDO)

        # ── Inserir miniatura ──────────────────────────────────────────────
        img_path = peca.get('img_path')
        if img_path:
            buf = miniatura_bytes(img_path)
            if buf:
                try:
                    xl_img = XLImage(buf)
                    xl_img.width  = 56
                    xl_img.height = 56
                    # Coluna 1 (A), linha row_idx
                    col_letra = get_column_letter(1)
                    ws.add_image(xl_img, f'{col_letra}{row_idx}')
                except Exception as e:
                    pass  # imagem com problema → célula vazia

        # Progresso
        if (row_idx - 1) % 50 == 0:
            print(f'  {row_idx - 1}/{total} peças processadas...')

    # ── Aba de instruções ──────────────────────────────────────────────────
    wi = wb.create_sheet('Instruções')
    wi.column_dimensions['A'].width = 60
    instrucoes = [
        ('COMO USAR ESTE FICHEIRO', True, '1B4B9C'),
        ('', False, '000000'),
        ('1. Não editar colunas ID, Ficheiro e Ano', False, '000000'),
        ('2. Preencher os campos de metadados', False, '000000'),
        ('3. Gravar o ficheiro Excel', False, '000000'),
        ('4. Correr: python excel_para_json.py', False, '000000'),
        ('', False, '000000'),
        ('PREÇO: 18,00 € ou 18€  (vazio = Sob consulta)', False, 'E8352A'),
        ('HASHTAGS válidas: cerâmica-única, cerâmica-molde, boneca,', False, 'E8352A'),
        ('  casa, arte, produto-alimentar, cestaria, outros', False, 'E8352A'),
    ]
    for i, (texto, bold, cor) in enumerate(instrucoes, 1):
        c = wi.cell(row=i, column=1, value=texto)
        c.font = Font(name='Arial', size=11, bold=bold, color=cor)

    wb.save(OUTPUT)
    print(f'\nExcel com imagens gerado: {OUTPUT}')
    print(f'Total: {total} peças')

def main():
    print('LandFeelings — Gerador de Excel com Miniaturas')
    print(f'Base: {BASE}\n')
    pecas = carregar_dados()
    if not pecas:
        print('Nenhuma peça encontrada.')
        return
    gerar_excel(pecas)
    print('\nPreenche o Excel e corre: python excel_para_json.py')

if __name__ == '__main__':
    main()
