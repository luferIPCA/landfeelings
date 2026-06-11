#!/usr/bin/env python3
"""
gerar_excel.py — LandFeelings
Gera ficheiro(s) Excel de metadados das peças dos catálogos.

Uso:
  python gerar_excel.py                     -> todos os anos, sem imagens
  python gerar_excel.py 2022                -> só o ano 2022, sem imagens
  python gerar_excel.py all --imagens       -> todos os anos, com miniaturas
  python gerar_excel.py 2022 --imagens      -> ano 2022, com miniaturas
  python gerar_excel.py all --tamanho=80    -> miniaturas de 80px

Ficheiros gerados:
  catalogo_metadados.xlsx            (all, sem imagens)
  catalogo_metadados_imagens.xlsx    (all, com imagens)
  catalogo_2022.xlsx                 (ano 2022, sem imagens)
  catalogo_2022_imagens.xlsx         (ano 2022, com imagens)
"""

import os, sys, json, io

BASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'images', 'Catalogos')

def instalar(pacote):
    import subprocess
    print(f'A instalar {pacote}...')
    subprocess.run([sys.executable, '-m', 'pip', 'install', pacote, '--user', '-q'])

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    instalar('openpyxl')
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

PIL_OK = False
try:
    from PIL import Image as PILImage
    PIL_OK = True
except ImportError:
    pass

# Cores
COR_CAB  = '1B4B9C'; COR_TXT = 'FFFFFF'; COR_ID = 'F0F4FF'
COR_ANO  = 'FFF8E8'; COR_LCK = 'F0F0F0'; COR_ALT = 'F9F9F9'; COR_BRD = 'CCCCCC'
lado   = Side(style='thin', color=COR_BRD)
bordas = Border(left=lado, right=lado, top=lado, bottom=lado)

HASHTAGS = ['cerâmica-única','cerâmica-molde','boneca','casa',
            'arte','produto-alimentar','cestaria','outros']

COLUNAS = [
    ('ID',           'id',        18, True,  'Identificador único — NÃO EDITAR'),
    ('Ficheiro',     'ficheiro',  25, True,  'Nome do ficheiro — NÃO EDITAR'),
    ('Ano',          'ano',       8,  True,  'Ano do catálogo — NÃO EDITAR'),
    ('Título',       'titulo',    28, False, 'Nome da peça (vazio = nome do ficheiro)'),
    ('Descrição',    'descricao', 40, False, 'Texto descritivo (aceita HTML)'),
    ('Autor',        'autor',     18, False, 'Artesã/autor da peça'),
    ('Preço',        'preco',     12, False, 'Ex: 18,00 €  (vazio = Sob consulta)'),
    ('Altura (cm)',  'altura',    11, False, 'Altura em centímetros'),
    ('Largura (cm)', 'largura',   11, False, 'Largura em centímetros'),
    ('Peso (g)',     'peso',      10, False, 'Peso em gramas'),
    ('HashTags',     'hashtags',  38, False, 'Separadas por vírgula: ' + ', '.join(HASHTAGS)),
]
COL_IMG = ('Img', 'img', 10, True, 'Miniatura')

def mk_cab(ws, row, col, txt, larg):
    c = ws.cell(row=row, column=col, value=txt)
    c.font = Font(bold=True, color=COR_TXT, name='Arial', size=10)
    c.fill = PatternFill('solid', fgColor=COR_CAB)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = bordas
    ws.column_dimensions[get_column_letter(col)].width = larg

def mk_cel(ws, row, col, val, locked, alt, campo):
    c = ws.cell(row=row, column=col, value=val)
    c.border = bordas
    c.font = Font(name='Arial', size=10, color='555555' if locked else '111111')
    c.alignment = Alignment(vertical='center', wrap_text=(col >= 4))
    if locked:
        cor = COR_ID if campo == 'id' else COR_ANO if campo == 'ano' else COR_LCK
        c.fill = PatternFill('solid', fgColor=cor)
    elif alt:
        c.fill = PatternFill('solid', fgColor=COR_ALT)

def mk_miniatura(img_path, sz):
    try:
        from PIL import Image as PILImage
        from openpyxl.drawing.image import Image as XLImage
        with PILImage.open(img_path) as img:
            img = img.convert('RGB')
            img.thumbnail((sz, sz), PILImage.LANCZOS)
            fundo = PILImage.new('RGB', (sz, sz), (255, 255, 255))
            fundo.paste(img, ((sz-img.width)//2, (sz-img.height)//2))
            buf = io.BytesIO()
            fundo.save(buf, format='PNG', optimize=True)
            buf.seek(0)
        xl = XLImage(buf)
        xl.width = sz - 6; xl.height = sz - 6
        return xl
    except Exception:
        return None

def hstr(peca):
    h = peca.get('hashtags', [])
    return ', '.join(h) if isinstance(h, list) else (h or '')

def carregar(ano):
    path = os.path.join(BASE, ano, 'index.json')
    if not os.path.exists(path):
        print(f'  {ano}: index.json não encontrado — corre primeiro gerar_catalogos.py')
        return []
    try:
        with open(path, 'r', encoding='utf-8') as f:
            dados = json.load(f)
    except Exception as e:
        print(f'  {ano}: erro — {e}'); return []
    pecas = []
    for i, p in enumerate(dados, 1):
        img = os.path.join(BASE, ano, p.get('ficheiro',''))
        pecas.append({
            'id': p.get('id', f'LF-{ano}-{i:03d}'), 'ficheiro': p.get('ficheiro',''),
            'ano': ano, 'titulo': p.get('titulo',''), 'descricao': p.get('descricao',''),
            'autor': p.get('autor',''), 'preco': p.get('preco',''),
            'altura': p.get('altura',''), 'largura': p.get('largura',''),
            'peso': p.get('peso',''), 'hashtags': hstr(p),
            'img_path': img if os.path.exists(img) else None,
        })
    print(f'  {ano}: {len(pecas)} peças'); return pecas

def gerar_aba(wb, pecas, ano, com_img, sz):
    ws = wb.create_sheet(title=str(ano))
    cols = ([COL_IMG] if com_img else []) + COLUNAS
    ws.freeze_panes = f'{"E" if com_img else "D"}2'
    ws.row_dimensions[1].height = 28
    for ci, (h, campo, larg, lock, _) in enumerate(cols, 1):
        mk_cab(ws, 1, ci, h, larg)
    for ri, p in enumerate(pecas, 2):
        ws.row_dimensions[ri].height = sz + 4 if com_img else 18
        for ci, (h, campo, larg, lock, _) in enumerate(cols, 1):
            if campo == 'img':
                if com_img and p.get('img_path'):
                    xl = mk_miniatura(p['img_path'], sz)
                    if xl: ws.add_image(xl, f'{get_column_letter(ci)}{ri}')
                continue
            mk_cel(ws, ri, ci, p.get(campo,''), lock, ri % 2 == 0, campo)
        if (ri - 1) % 100 == 0 and ri > 2:
            print(f'    {ri-1}/{len(pecas)}...')
    print(f'  Aba "{ano}": {len(pecas)} peças')

def instrucoes(wb):
    wi = wb.create_sheet('Instruções')
    wi.column_dimensions['A'].width = 70
    linhas = [
        ('COMO USAR', True, '1B4B9C'), ('', False, '000000'),
        ('1. NÃO editar colunas ID, Ficheiro e Ano', False, '333333'),
        ('2. Preencher os metadados nas restantes colunas', False, '333333'),
        ('3. Gravar o ficheiro Excel', False, '333333'),
        ('4. Correr: python importar_excel.py', False, '333333'),
        ('', False, '000000'),
        ('IMPORTAÇÃO', True, 'E8352A'),
        ('  python importar_excel.py                          -> importa tudo', False, '333333'),
        ('  python importar_excel.py 2022                     -> só o ano 2022', False, '333333'),
        ('  python importar_excel.py --ficheiro catalogo_2022.xlsx  -> ficheiro específico', False, '333333'),
        ('', False, '000000'),
        ('PREÇO: 18,00 €  ou  18€  (vazio = Sob consulta)', False, 'E8352A'),
        ('HASHTAGS válidas (separar por vírgula):', True, 'E8352A'),
    ] + [(f'  • {h}', False, '555555') for h in HASHTAGS]
    for i, (txt, bold, cor) in enumerate(linhas, 1):
        c = wi.cell(row=i, column=1, value=txt)
        c.font = Font(name='Arial', size=11, bold=bold, color=cor)

def main():
    args = sys.argv[1:]
    com_img = '--imagens' in args
    args = [a for a in args if a != '--imagens']
    sz = 70
    for a in args:
        if a.startswith('--tamanho='):
            try: sz = int(a.split('=')[1])
            except: pass
    args = [a for a in args if not a.startswith('--tamanho')]
    ano_param = args[0].lower() if args else 'all'

    if com_img and not PIL_OK:
        instalar('Pillow')
        try:
            from PIL import Image as PILImage
        except:
            print('Erro ao instalar Pillow. Corre: pip install Pillow'); sys.exit(1)

    if not os.path.exists(BASE):
        print(f'Pasta não encontrada: {BASE}'); sys.exit(1)

    disponiveis = sorted([d for d in os.listdir(BASE) if os.path.isdir(os.path.join(BASE, d))])

    if ano_param == 'all':
        anos = disponiveis
    elif ano_param in disponiveis:
        anos = [ano_param]
    else:
        print(f'Ano "{ano_param}" não encontrado. Disponíveis: {", ".join(disponiveis)}'); sys.exit(1)

    suf = '_imagens' if com_img else ''
    nome = f'catalogo_metadados{suf}.xlsx' if ano_param == 'all' else f'catalogo_{ano_param}{suf}.xlsx'
    output = os.path.join(os.path.dirname(os.path.abspath(__file__)), nome)

    print(f'LandFeelings — Gerador de Excel')
    print(f'Anos: {", ".join(anos)}  |  Imagens: {"sim (" + str(sz) + "px)" if com_img else "não"}')
    print(f'Output: {nome}\n')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for ano in anos:
        print(f'A processar {ano}...')
        pecas = carregar(ano)
        if pecas:
            gerar_aba(wb, pecas, ano, com_img, sz)
    instrucoes(wb)
    wb.save(output)
    print(f'\nExcel gerado: {output}')
    print('Preenche os metadados e corre: python importar_excel.py')

if __name__ == '__main__':
    main()
