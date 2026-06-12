#!/usr/bin/env python3
"""
importar_excel.py — LandFeelings
Importa metadados do Excel e actualiza os index.json dos catálogos.
Usa o campo ID como chave — nunca perde metadados já preenchidos.
Suporta adição de novas peças (novas linhas no Excel após gerar_catalogos.py).

Uso:
  python importar_excel.py                                    -> lê catalogo_metadados.xlsx, importa tudo
  python importar_excel.py 2022                               -> importa só o ano 2022
  python importar_excel.py --ficheiro catalogo_2022.xlsx      -> ficheiro específico, tudo
  python importar_excel.py 2022 --ficheiro catalogo_2022.xlsx -> ficheiro específico, ano 2022
  python importar_excel.py --ficheiro catalogo_metadados_imagens.xlsx -> versão com imagens
"""

import os, sys, json

BASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'images', 'Catalogos')

def instalar(pacote):
    import subprocess
    print(f'A instalar {pacote}...')
    subprocess.run([sys.executable, '-m', 'pip', 'install', pacote, '--user', '-q'])

try:
    import openpyxl
except ImportError:
    instalar('openpyxl')
    import openpyxl

# Campos que podem ser importados do Excel (na ordem das colunas sem imagem)
CAMPOS_EDITAVEIS = ['titulo', 'descricao', 'autor', 'preco',
                    'altura', 'largura', 'peso', 'hashtags']
CAMPOS_BLOQUEADOS = ['id', 'ficheiro', 'ano']
TODOS_CAMPOS = CAMPOS_BLOQUEADOS + CAMPOS_EDITAVEIS

def ler_excel(ficheiro, ano_filtro=None):
    """
    Lê o Excel e devolve dicionário {id: dados} e {ano: [ids]}.
    Suporta Excel com ou sem coluna de imagem (detecta automaticamente).
    Suporta Excel com uma aba por ano ou tudo numa só aba.
    """
    if not os.path.exists(ficheiro):
        print(f'Ficheiro não encontrado: {ficheiro}')
        print('Corre primeiro: python gerar_excel.py')
        return {}, {}

    wb = openpyxl.load_workbook(ficheiro, read_only=True, data_only=True)
    dados = {}
    por_ano = {}

    for nome_aba in wb.sheetnames:
        if nome_aba == 'Instruções':
            continue
        # Se filtro de ano activo, só processar aba correspondente
        if ano_filtro and nome_aba != ano_filtro:
            continue

        ws = wb[nome_aba]
        cabecalho = []
        for row in ws.iter_rows(values_only=True):
            if not cabecalho:
                # Primeira linha = cabeçalho — detectar colunas
                cabecalho = [str(c).strip() if c else '' for c in row]
                # Mapear nomes de colunas para campos JSON
                mapa = {}
                nomes_para_campos = {
                    'ID': 'id', 'Ficheiro': 'ficheiro', 'Ano': 'ano',
                    'Título': 'titulo', 'Descrição': 'descricao', 'Autor': 'autor',
                    'Preço': 'preco', 'Altura (cm)': 'altura', 'Largura (cm)': 'largura',
                    'Peso (g)': 'peso', 'HashTags': 'hashtags', 'Img': None
                }
                for i, nome in enumerate(cabecalho):
                    campo = nomes_para_campos.get(nome)
                    if campo is not None and campo != 'img':
                        mapa[i] = campo
                continue

            # Linha vazia — verificar se alguma célula tem dados (não só a primeira)
            if not any(v for v in row):
                continue

            peca = {}
            for i, val in enumerate(row):
                campo = mapa.get(i)
                if campo:
                    peca[campo] = str(val).strip() if val is not None else ''

            pid = peca.get('id', '')
            ano = peca.get('ano', nome_aba)
            if not pid:
                continue

            # Converter hashtags para lista
            h = peca.get('hashtags', '')
            peca['hashtags'] = [x.strip() for x in h.split(',') if x.strip()] if h else []

            dados[pid] = peca
            por_ano.setdefault(ano, []).append(pid)

    wb.close()
    print(f'Excel lido: {len(dados)} peças em {len(por_ano)} ano(s)')
    return dados, por_ano

def actualizar_json(dados_excel, por_ano, ano_filtro=None):
    """Actualiza os index.json com os dados do Excel."""
    if not os.path.exists(BASE):
        print(f'Pasta não encontrada: {BASE}'); return

    anos = sorted([d for d in os.listdir(BASE) if os.path.isdir(os.path.join(BASE, d))])
    total = 0

    for ano in anos:
        if ano_filtro and ano != ano_filtro:
            continue
        if ano not in por_ano and not any(p.get('ano') == ano for p in dados_excel.values()):
            print(f'  {ano}: sem dados no Excel — ignorado')
            continue

        index_path = os.path.join(BASE, ano, 'index.json')
        if not os.path.exists(index_path):
            print(f'  {ano}: index.json não encontrado'); continue

        try:
            with open(index_path, 'r', encoding='utf-8') as f:
                pecas = json.load(f)
        except Exception as e:
            print(f'  {ano}: erro ao ler JSON — {e}'); continue

        # Mapear IDs existentes no JSON
        ids_existentes = {p.get('id'): i for i, p in enumerate(pecas) if p.get('id')}
        actualizadas = 0
        novas = 0

        # IDs do Excel para este ano
        ids_excel_ano = [pid for pid, p in dados_excel.items() if p.get('ano') == ano]

        for pid in ids_excel_ano:
            dados = dados_excel[pid]
            if pid in ids_existentes:
                # Actualizar entrada existente
                idx = ids_existentes[pid]
                for campo in CAMPOS_EDITAVEIS:
                    if campo in dados:
                        pecas[idx][campo] = dados[campo]
                actualizadas += 1
            else:
                # Nova peça — adicionar ao JSON
                pecas.append({
                    'id':        dados.get('id', ''),
                    'ficheiro':  dados.get('ficheiro', ''),
                    'titulo':    dados.get('titulo', ''),
                    'descricao': dados.get('descricao', ''),
                    'autor':     dados.get('autor', ''),
                    'preco':     dados.get('preco', ''),
                    'altura':    dados.get('altura', ''),
                    'largura':   dados.get('largura', ''),
                    'peso':      dados.get('peso', ''),
                    'hashtags':  dados.get('hashtags', []),
                })
                novas += 1

        with open(index_path, 'w', encoding='utf-8') as f:
            json.dump(pecas, f, ensure_ascii=False, indent=2)

        msg = f'{actualizadas} actualizadas'
        if novas: msg += f', {novas} novas'
        print(f'  {ano}: {msg} → index.json guardado')
        total += actualizadas + novas

    print(f'\nConcluído! {total} peças processadas.')

def main():
    args = sys.argv[1:]

    # Extrair --ficheiro
    ficheiro = None
    for i, a in enumerate(args):
        if a.startswith('--ficheiro='):
            ficheiro = a.split('=', 1)[1]
            args = args[:i] + args[i+1:]
            break
        elif a == '--ficheiro' and i + 1 < len(args):
            ficheiro = args[i+1]
            args = args[:i] + args[i+2:]
            break

    # Extrair ano
    ano_filtro = None
    for a in args:
        if not a.startswith('--'):
            ano_filtro = a if a.lower() != 'all' else None
            break

    # Ficheiro por defeito
    if not ficheiro:
        ficheiro = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                'catalogo_metadados.xlsx')
    elif not os.path.isabs(ficheiro):
        ficheiro = os.path.join(os.path.dirname(os.path.abspath(__file__)), ficheiro)

    print(f'LandFeelings — Importador Excel → JSON')
    print(f'Ficheiro: {os.path.basename(ficheiro)}')
    print(f'Ano: {ano_filtro or "todos"}\n')

    dados, por_ano = ler_excel(ficheiro, ano_filtro)
    if not dados:
        print('Nenhum dado encontrado no Excel.'); return

    actualizar_json(dados, por_ano, ano_filtro)
    print('\nRecarrega o browser para ver as alterações.')

if __name__ == '__main__':
    main()
